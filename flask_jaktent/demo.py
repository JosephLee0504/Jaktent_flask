import sqlite3
def init_db(dbpath) :

    # sql1 = '''
    #     create table Jaktent
    #     (
    #     Name text primary key,
    #     type text,
    #     Prensenters text,
    #     startDate text,
    #     endDate text ,
    #     startTime text,
    #     endTime text,
    #     content text,
    #     URL text
    #     )'''
    #
    sql2 = '''
        create table JaktentSpeaker
        ( 
        Name text primary key,
        Career text,
        Description text,
        URL text
        )'''

    # Create data table
    conn = sqlite3.connect(dbpath)
    cursor = conn.cursor()
    cursor.execute(sql2)
    conn.commit()
    conn.close()

dbpath = "Jaktent.db"



def movie():
    datalist = []
    a = "Film"
    #print(type(a))
    conn = sqlite3.connect("Jaktent.db")
    cur = conn.cursor()
    sql = "select * from JaktentEvent  "
    data = cur.execute(sql)
    print(data)
    for item in data:
        #print(item[0])
        #print(type(item[0]))
        if a in item[0]:
            #print(item[0])
            datalist.append(item)
            print(item)
        #datalist.append(item)
    cur.close()
    conn.close()

def rem():
    datalist = []
    rdata = []
    pdata = []
    a = input("input：")
    conn = sqlite3.connect("Jaktent.db")
    cur = conn.cursor()

    sql2 = "select * from JaktentEvent"
    data2 = cur.execute(sql2)
    for item in data2:
        if a in item[0]:
            rdata.append(item)
    num = rdata[0][9]


    sql1 = "select id2 from similar where number > 0.2 and id = ?"
    v = (num,)
    data1 = cur.execute(sql1,v)
    for i in data1:
        datalist.append(i[0])
    print(datalist)

    for item in datalist:
        sql3 = "select * from JaktentEvent where id = ? "
        v = (item,)
        data3 = cur.execute(sql3,v)
        for a in data3:
            pdata.append(a)

    print(pdata)
    cur.close()
    conn.close()
import xlwt
import openpyxl as p

def b():
    wb = p.load_workbook('question_record.xlsx')
    #ws = wb.get_sheet_by_name('Sheet1')
    ws = wb['Sheet1']
    print(ws.max_column)
    print(ws.max_row)





def read():
    workbook = xlwt.Workbook(encoding = "utf-8")
    sheet = workbook.add_sheet('Sheet1')

    a = input("input")
    b = input("output")
    c = (a,b)

    wb = p.load_workbook('question_record.xlsx')
    #ws = wb.get_sheet_by_name('Sheet1')
    ws = wb['Sheet1']
    print(ws.max_column)
    d = ws.max_row+1

    for j in range(0,2):
        sheet.write(d,j,c[j])

    workbook.save('question_record.xlsx')

def re():
    conn = sqlite3.connect("Jaktent.db")
    cur = conn.cursor()
    sql = "insert into recording(question,recording) values(?,?) "
    a = input("a")
    b = input("b")
    cur.execute(sql,(a,b))
    conn.commit()
    cur.close()
    conn.close()
re()